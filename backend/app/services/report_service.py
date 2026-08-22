from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from sqlalchemy.orm import Session

from app.exceptions.common import (
    NotFoundException,
    BadRequestException,
)

from app.exceptions.messages import FORM_NOT_FOUND

from app.repositories.report_repository import ReportRepository

from app.schemas.report import (
    ReportFilter,
    ResponseStatisticsResponse,
    FormResponseStatistics,
    FieldOptionAnalytics,
    FieldAnalytics,
    FormAnalyticsResponse,
    ResponseTrendItem,
    ResponseTrendResponse,
)


class ReportService:

    def __init__(self) -> None:
        self.report_repository = ReportRepository()

    # =========================================================
    # RESPONSE STATISTICS
    # =========================================================

    def get_response_statistics(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> ResponseStatisticsResponse:

        self._validate_date_range(
            filters.from_date,
            filters.to_date,
        )

        # -----------------------------------------------------
        # Validate form
        # -----------------------------------------------------

        if filters.form_id is not None:
            form = self.report_repository.get_form_by_id(
                db,
                filters.form_id,
            )

            if form is None:
                raise NotFoundException(
                    FORM_NOT_FOUND
                )

        # -----------------------------------------------------
        # Total responses
        # -----------------------------------------------------

        total_responses = (
            self.report_repository.get_total_responses(
                db,
                form_id=filters.form_id,
                user_id=filters.user_id,
                from_date=filters.from_date,
                to_date=filters.to_date,
            )
        )

        # -----------------------------------------------------
        # Current time
        # -----------------------------------------------------

        now = datetime.now(timezone.utc)

        # -----------------------------------------------------
        # Today
        # -----------------------------------------------------

        start_of_today = datetime(
            now.year,
            now.month,
            now.day,
            tzinfo=timezone.utc,
        )

        responses_today = (
            self.report_repository.get_total_responses(
                db,
                form_id=filters.form_id,
                user_id=filters.user_id,
                from_date=start_of_today,
                to_date=now,
            )
        )

        # -----------------------------------------------------
        # This week
        # -----------------------------------------------------

        start_of_week = (
            start_of_today
            - timedelta(
                days=start_of_today.weekday()
            )
        )

        responses_this_week = (
            self.report_repository.get_total_responses(
                db,
                form_id=filters.form_id,
                user_id=filters.user_id,
                from_date=start_of_week,
                to_date=now,
            )
        )

        # -----------------------------------------------------
        # This month
        # -----------------------------------------------------

        start_of_month = datetime(
            now.year,
            now.month,
            1,
            tzinfo=timezone.utc,
        )

        responses_this_month = (
            self.report_repository.get_total_responses(
                db,
                form_id=filters.form_id,
                user_id=filters.user_id,
                from_date=start_of_month,
                to_date=now,
            )
        )

        return ResponseStatisticsResponse(
            total_responses=total_responses,
            responses_today=responses_today,
            responses_this_week=responses_this_week,
            responses_this_month=responses_this_month,
        )

    # =========================================================
    # FORM-WISE RESPONSE STATISTICS
    # =========================================================

    def get_form_response_statistics(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> list[FormResponseStatistics]:

        self._validate_date_range(
            filters.from_date,
            filters.to_date,
        )

        if filters.form_id is not None:

            form = self.report_repository.get_form_by_id(
                db,
                filters.form_id,
            )

            if form is None:
                raise NotFoundException(
                    FORM_NOT_FOUND
                )

        rows = (
            self.report_repository.get_form_response_statistics(
                db,
                form_id=filters.form_id,
                from_date=filters.from_date,
                to_date=filters.to_date,
            )
        )

        return [
            FormResponseStatistics(
                form_id=form_id,
                form_title=form_title,
                total_responses=total_responses,
            )
            for form_id, form_title, total_responses in rows
        ]

    # =========================================================
    # FORM-WISE ANALYTICS
    # =========================================================

    def get_form_analytics(
        self,
        db: Session,
        form_id: int,
        user_id: int | None = None,
        from_date: datetime | None = None,
        to_date: datetime | None = None,
    ) -> FormAnalyticsResponse:

        self._validate_date_range(
            from_date,
            to_date,
        )

        # -----------------------------------------------------
        # Get form
        # -----------------------------------------------------

        form = self.report_repository.get_form_by_id(
            db,
            form_id,
        )

        if form is None:
            raise NotFoundException(
                FORM_NOT_FOUND
            )

        # -----------------------------------------------------
        # Total responses
        # -----------------------------------------------------

        total_responses = (
            self.report_repository.get_total_responses(
                db,
                form_id=form_id,
                user_id=user_id,
                from_date=from_date,
                to_date=to_date,
            )
        )

        # -----------------------------------------------------
        # Get fields
        # -----------------------------------------------------

        fields = self.report_repository.get_form_fields(
            db,
            form_id,
        )

        field_analytics: list[FieldAnalytics] = []

        # -----------------------------------------------------
        # Analyze every field
        # -----------------------------------------------------

        for field in fields:

            values = (
                self.report_repository.get_field_response_values(
                    db,
                    field_id=field.id,
                    form_id=form_id,
                    user_id=user_id,
                    from_date=from_date,
                    to_date=to_date,
                )
            )

            # -------------------------------------------------
            # Count values
            # -------------------------------------------------

            option_counts: dict[str, int] = {}

            for value in values:

                option_counts[value] = (
                    option_counts.get(value, 0) + 1
                )

            options = [
                FieldOptionAnalytics(
                    option=option,
                    count=count,
                )
                for option, count in option_counts.items()
            ]

            # -------------------------------------------------
            # Field type
            # -------------------------------------------------

            field_type = field.field_type

            if hasattr(field_type, "value"):
                field_type_value = str(
                    field_type.value
                )
            else:
                field_type_value = str(
                    field_type
                )

            field_analytics.append(
                FieldAnalytics(
                    field_id=field.id,
                    field_label=field.label,
                    field_type=field_type_value,
                    total_responses=len(values),
                    options=options,
                )
            )

        return FormAnalyticsResponse(
            form_id=form.id,
            form_title=form.title,
            total_responses=total_responses,
            fields=field_analytics,
        )

    # =========================================================
    # RESPONSE TREND
    # =========================================================

    def get_response_trend(
        self,
        db: Session,
        from_date: datetime,
        to_date: datetime,
        form_id: int | None = None,
    ) -> ResponseTrendResponse:

        self._validate_date_range(
            from_date,
            to_date,
        )

        # -----------------------------------------------------
        # Validate form
        # -----------------------------------------------------

        if form_id is not None:

            form = self.report_repository.get_form_by_id(
                db,
                form_id,
            )

            if form is None:
                raise NotFoundException(
                    FORM_NOT_FOUND
                )

        rows = self.report_repository.get_response_trend(
            db,
            from_date=from_date,
            to_date=to_date,
            form_id=form_id,
        )

        data = [
            ResponseTrendItem(
                date=str(date),
                count=count,
            )
            for date, count in rows
        ]

        return ResponseTrendResponse(
            form_id=form_id,
            from_date=from_date,
            to_date=to_date,
            data=data,
        )

    # =========================================================
    # PREPARE EXPORT DATA
    # =========================================================

    def get_export_data(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> list[dict[str, str]]:

        self._validate_date_range(
            filters.from_date,
            filters.to_date,
        )

        # -----------------------------------------------------
        # Validate form
        # -----------------------------------------------------

        if filters.form_id is not None:

            form = self.report_repository.get_form_by_id(
                db,
                filters.form_id,
            )

            if form is None:
                raise NotFoundException(
                    FORM_NOT_FOUND
                )

        responses = (
            self.report_repository.get_responses_for_export(
                db,
                form_id=filters.form_id,
                user_id=filters.user_id,
                from_date=filters.from_date,
                to_date=filters.to_date,
            )
        )

        export_data: list[dict[str, str]] = []

        # -----------------------------------------------------
        # Cache field labels
        # -----------------------------------------------------

        form_fields_cache: dict[
            int,
            dict[int, str]
        ] = {}

        for response in responses:

            if response.form_id not in form_fields_cache:

                fields = (
                    self.report_repository.get_form_fields(
                        db,
                        response.form_id,
                    )
                )

                form_fields_cache[
                    response.form_id
                ] = {
                    field.id: field.label
                    for field in fields
                }

            field_map = form_fields_cache[
                response.form_id
            ]

            # -------------------------------------------------
            # User
            # -------------------------------------------------

            user_email = ""

            if response.user_id is not None:

                user = (
                    self.report_repository.get_user_by_id(
                        db,
                        response.user_id,
                    )
                )

                if user is not None:
                    user_email = str(
                        user.email
                    )

            # -------------------------------------------------
            # Base row
            # -------------------------------------------------

            row: dict[str, str] = {
                "Response ID": str(
                    response.id
                ),
                "Form ID": str(
                    response.form_id
                ),
                "User ID": (
                    str(response.user_id)
                    if response.user_id is not None
                    else ""
                ),
                "User": user_email,
                "Submitted At": (
                    response.submitted_at.isoformat()
                    if response.submitted_at is not None
                    else ""
                ),
            }

            # -------------------------------------------------
            # Response details
            # -------------------------------------------------

            details = (
                self.report_repository.get_response_details(
                    db,
                    response.id,
                )
            )

            for detail in details:

                field_label = field_map.get(
                    detail.field_id
                )

                if field_label is not None:

                    row[field_label] = (
                        detail.value or ""
                    )

            export_data.append(row)

        return export_data

    # =========================================================
    # EXPORT EXCEL
    # =========================================================

    def generate_excel_report(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> BytesIO:

        data = self.get_export_data(
            db=db,
            filters=filters,
        )

        workbook = Workbook()

        worksheet = workbook.active

        if worksheet is None:
            raise RuntimeError("Failed to create Excel worksheet.")

        worksheet.title = "Responses"

        # =========================================================
        # EMPTY REPORT
        # =========================================================

        if not data:
            worksheet["A1"] = "No response data found."

            worksheet["A1"].font = Font(
                bold=True
            )

        else:
            # =====================================================
            # COLLECT ALL COLUMNS
            # =====================================================

            columns: list[str] = []

            for row in data:
                for key in row.keys():
                    if key not in columns:
                        columns.append(key)

            # =====================================================
            # HEADER
            # =====================================================

            for column_index, column_name in enumerate(
                columns,
                start=1,
            ):
                cell = worksheet.cell(
                    row=1,
                    column=column_index,
                    value=column_name,
                )

                cell.font = Font(
                    bold=True
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9EAF7",
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            # =====================================================
            # DATA
            # =====================================================

            for row_index, row_data in enumerate(
                data,
                start=2,
            ):
                for column_index, column_name in enumerate(
                    columns,
                    start=1,
                ):
                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=row_data.get(
                            column_name,
                            "",
                        ),
                    )

            # =====================================================
            # AUTO COLUMN WIDTH
            # =====================================================

            for column_index, column_name in enumerate(
                columns,
                start=1,
            ):
                max_length = len(column_name)

                for row_index in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    value = worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).value

                    if value is not None:
                        max_length = max(
                            max_length,
                            len(str(value)),
                        )

                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = min(
                    max_length + 2,
                    40,
                )

        # =========================================================
        # FREEZE HEADER
        # =========================================================

        worksheet.freeze_panes = "A2"

        # =========================================================
        # FILTER
        # =========================================================

        if data:
            worksheet.auto_filter.ref = worksheet.dimensions

        # =========================================================
        # SAVE TO MEMORY
        # =========================================================

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output

    # =========================================================
    # EXPORT PDF
    # =========================================================

    def generate_pdf_report(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> BytesIO:

        data = self.get_export_data(
            db,
            filters,
        )

        output = BytesIO()

        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )

        styles = getSampleStyleSheet()

        elements = []

        # -----------------------------------------------------
        # Title
        # -----------------------------------------------------

        elements.append(
            Paragraph(
                "Form Response Report",
                styles["Title"],
            )
        )

        elements.append(
            Spacer(
                1,
                5 * mm,
            )
        )

        # -----------------------------------------------------
        # Filter information
        # -----------------------------------------------------

        filter_text = []

        if filters.form_id is not None:
            filter_text.append(
                f"Form ID: {filters.form_id}"
            )

        if filters.user_id is not None:
            filter_text.append(
                f"User ID: {filters.user_id}"
            )

        if filters.from_date is not None:
            filter_text.append(
                "From: "
                + filters.from_date.isoformat()
            )

        if filters.to_date is not None:
            filter_text.append(
                "To: "
                + filters.to_date.isoformat()
            )

        if filter_text:

            elements.append(
                Paragraph(
                    " | ".join(filter_text),
                    styles["Normal"],
                )
            )

            elements.append(
                Spacer(
                    1,
                    5 * mm,
                )
            )

        # -----------------------------------------------------
        # No data
        # -----------------------------------------------------

        if not data:

            elements.append(
                Paragraph(
                    "No response data found.",
                    styles["Normal"],
                )
            )

        else:

            # -------------------------------------------------
            # Columns
            # -------------------------------------------------

            columns: list[str] = []

            for row in data:

                for key in row.keys():

                    if key not in columns:
                        columns.append(key)

            table_data = [
                columns
            ]

            # -------------------------------------------------
            # Rows
            # -------------------------------------------------

            for row in data:

                table_data.append(
                    [
                        str(
                            row.get(
                                column,
                                "",
                            )
                        )
                        for column in columns
                    ]
                )

            # -------------------------------------------------
            # Calculate column widths
            # -------------------------------------------------

            available_width = (
                landscape(A4)[0]
                - 20 * mm
            )

            column_width = (
                available_width
                / len(columns)
            )

            table = Table(
                table_data,
                repeatRows=1,
                colWidths=[
                    column_width
                    for _ in columns
                ],
            )

            # -------------------------------------------------
            # Table styling
            # -------------------------------------------------

            table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor(
                                "#D9EAF7"
                            ),
                        ),
                        (
                            "TEXTCOLOR",
                            (0, 0),
                            (-1, 0),
                            colors.black,
                        ),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, 0),
                            "Helvetica-Bold",
                        ),
                        (
                            "FONTSIZE",
                            (0, 0),
                            (-1, -1),
                            7,
                        ),
                        (
                            "GRID",
                            (0, 0),
                            (-1, -1),
                            0.5,
                            colors.grey,
                        ),
                        (
                            "VALIGN",
                            (0, 0),
                            (-1, -1),
                            "TOP",
                        ),
                        (
                            "ALIGN",
                            (0, 0),
                            (-1, -1),
                            "LEFT",
                        ),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [
                                colors.white,
                                colors.HexColor(
                                    "#F7F7F7"
                                ),
                            ],
                        ),
                    ]
                )
            )

            elements.append(table)

        # -----------------------------------------------------
        # Build PDF
        # -----------------------------------------------------

        document.build(elements)

        output.seek(0)

        return output

    # =========================================================
    # VALIDATION
    # =========================================================

    def _validate_date_range(
        self,
        from_date: datetime | None,
        to_date: datetime | None,
    ) -> None:

        if (
            from_date is not None
            and to_date is not None
            and from_date > to_date
        ):

            raise BadRequestException(
                "from_date cannot be greater than to_date."
            )